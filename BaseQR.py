from warnings import catch_warnings

import numpy as np
from openpyxl import Workbook
from openpyxl.styles import PatternFill
from pycparser.c_ast import Break

from NumeroBinario import CadenaBinario, NumBinario, LimpiarString
from PintarExcel import LongitudInput

fila = range(1,27)
abecedario = ["A", "B", "C", "D", "E", "F", "G", "H", "I", "J", "K", "L", "M", "N", "O", "P", "Q", "R", "S", "T",
                  "U", "V", "W", "X", "Y", "Z", "AA"]
def PintarBase(Color, Rango):
    fill = PatternFill(start_color=Color, end_color=Color, fill_type="solid")
    for i in range(1, Rango):
        for i2 in range(1, Rango):
            celda = abecedario[i - 1]  + str(i2)
            ws[celda].fill = fill
            ws.column_dimensions[abecedario[i - 1]].width = 3

def PintarRango(Color, Rango, Linea, PuntoInicio,Columna, Fila ):
    fill = PatternFill(start_color=Color, end_color=Color, fill_type="solid")
    if Linea == "Fila":
        for i in range(1, Rango):
            celda = abecedario[i + PuntoInicio - 1] + str(Fila)
            ws[celda].fill = fill
            ws.column_dimensions[abecedario[i + PuntoInicio - 1]].width = 3
    else:
        for i in range(1, Rango):
            celda = Columna + str(i + PuntoInicio)
            ws[celda].fill = fill
            ws.column_dimensions[abecedario[i - 1]].width = 3

def PintarCelda(Celda, Color):
    fill = PatternFill(start_color=Color, end_color=Color, fill_type="solid")
    ws[Celda].fill = fill

def AlternarPintar(ColorQR, ColorFondo, Rango, Linea, PuntoInicio, Columna, Fila, NSaltos ):
    fill = PatternFill(start_color=ColorQR, end_color=ColorQR, fill_type="solid")
    if Linea == "Fila":
        for aux in range(0, Rango , NSaltos):
            try:
                PintarCelda(abecedario[aux + PuntoInicio] + str(Fila), ColorQR)

                aux = aux + PuntoInicio
                print("El contador i es " + str(aux))

            except: print("list index out of range")


    else:
        for aux in range(0, Rango, NSaltos):
            try:
                PintarCelda(Columna+ str(aux + PuntoInicio), ColorQR)

                aux = aux + PuntoInicio

            except:
                print("list index out of range")


def PintarBaseQR(ColorQR):
    PintarRango(ColorQR, 8, "Fila", 1, "A", 2)
    PintarRango(ColorQR, 8, "Columna", 1, "B", 2)
    PintarRango(ColorQR, 8, "Fila", 1, "A", 8)
    PintarRango(ColorQR, 8, "Columna", 19, "H", 20)
    PintarRango(ColorQR, 8, "Fila", 1, "A", 20)
    PintarRango(ColorQR, 8, "Columna", 19, "B", 26)
    PintarRango(ColorQR, 8, "Fila", 1, "A", 26)
    PintarRango(ColorQR, 8, "Columna", 1, "H", 2)
    PintarRango(ColorQR, 8, "Columna", 1, "T", 20)
    PintarRango(ColorQR, 8, "Fila", 19, "T", 2)
    PintarRango(ColorQR, 8, "Columna", 1, "Z", 26)
    PintarRango(ColorQR, 8, "Fila", 19, "Z", 8)
    PintarRango(ColorQR, 4, "Fila", 3, "A", 4)
    PintarRango(ColorQR, 4, "Fila", 3, "A", 5)
    PintarRango(ColorQR, 4, "Fila", 3, "A", 6)
    PintarRango(ColorQR, 4, "Fila", 3, "A", 22)
    PintarRango(ColorQR, 4, "Fila", 3, "A", 23)
    PintarRango(ColorQR, 4, "Fila", 3, "A", 24)
    PintarRango(ColorQR, 4, "Fila", 21, "A", 4)
    PintarRango(ColorQR, 4, "Fila", 21, "A", 5)
    PintarRango(ColorQR, 4, "Fila", 21, "A", 6)
    PintarRango(ColorQR, 6, "Fila", 17, "A", 18)
    PintarRango(ColorQR, 6, "Columna", 17, "R", 2)
    PintarRango(ColorQR, 6, "Fila", 17, "A", 22)
    PintarRango(ColorQR, 6, "Columna", 17, "V", 20)
    PintarCelda("T20", ColorQR)
    PintarCelda("J19", ColorQR)


def PintarDiagonal(Fila, Columna, ColorQR, ColorFondo, Cadena):
    print("Empiezo a pintar diagonal")
    print(f"Longitud de la cadena es: {len(Cadena)} ")
    for i in range (0,len(Cadena)):
        print(f"Cadena {Cadena} con índice {i} es {Cadena[i]} ")
        if (Cadena[i] == "0"):
            Diagonal(ColorFondo, i, Columna, Fila)
        else:
            Diagonal(ColorQR, i , Columna, Fila)


def Diagonal(Color, Contador, Columna, Fila):
    if Contador == 0:
        Celda = abecedario[Columna] + str(Fila)
        PintarCelda(Celda, Color)
    elif Contador == 1:
        Celda = abecedario[Columna - 1] + str(Fila)
        PintarCelda(Celda, Color)
    elif Contador == 2:
        Celda = abecedario[Columna] + str(Fila - 1)
        PintarCelda(Celda, Color)
    elif Contador == 3:
        Celda = abecedario[Columna - 1] + str(Fila - 1)
        PintarCelda(Celda, Color)
    elif Contador == 4:
        Celda = abecedario[Columna] + str(Fila - 2)
        PintarCelda(Celda, Color)
    elif Contador == 5:
        Celda = abecedario[Columna - 1] + str(Fila - 2)
        PintarCelda(Celda, Color)
    elif Contador == 6:
        Celda = abecedario[Columna] + str(Fila - 3)
        PintarCelda(Celda, Color)
    elif Contador == 7:
        Celda = abecedario[Columna - 1] + str(Fila - 3)
        PintarCelda(Celda, Color)


def CeldasOcupadasCuadrado(inicioFila,inicioColumna, Rango):
    CeldasTomadas = []
    for Fila in (range(inicioFila, Rango + inicioFila)):
        for Columna in (range(inicioColumna, Rango + inicioColumna)):
            CeldasTomadas.append(abecedario[Columna] + str(Fila + 1))
    return CeldasTomadas

def CeldasOcupadasSolo(inicioFila,inicioColumna, Rango, Boolean):
    CeldasTomadas = []
    if Boolean:
        for Fila in (range(inicioFila, Rango + inicioFila)):
            CeldasTomadas.append(abecedario[inicioColumna] + str(Fila + 1))
    else:
        for Columna in (range(inicioColumna, Rango + inicioColumna)):
            CeldasTomadas.append(abecedario[Columna] + str(inicioFila + 1))

    return CeldasTomadas


print(CeldasOcupadasCuadrado(1,1, 9)) #ARRIBA IZQUIERDA
print(CeldasOcupadasCuadrado(18,1, 8)) #ABAJO IZQUIERDA
print(CeldasOcupadasSolo(18,10, 8, True)) #ABAJO IZQUIERDA
print(CeldasOcupadasSolo(9,18, 8, False))
print(CeldasOcupadasCuadrado(1,18, 8)) # ARRIBA DERECHA
print(CeldasOcupadasCuadrado(17,17, 5))  #CUADRADO SOLO
print(CeldasOcupadasCuadrado(6,11, 1))
print(CeldasOcupadasCuadrado(6,13, 1))
print(CeldasOcupadasCuadrado(6,15, 1))
print(CeldasOcupadasCuadrado(6,17, 1))
print(CeldasOcupadasCuadrado(11,6, 1))
print(CeldasOcupadasCuadrado(13,6, 1))
print(CeldasOcupadasCuadrado(15,6, 1))
print(CeldasOcupadasCuadrado(17,6, 1))


TotalCeldasOcupadas = (CeldasOcupadasCuadrado(6,11, 1) +  CeldasOcupadasCuadrado(6,13, 1) +  CeldasOcupadasCuadrado(6,15, 1) + CeldasOcupadasCuadrado(6,17, 1) + CeldasOcupadasCuadrado(11,6, 1) +  CeldasOcupadasCuadrado(13,6, 1) +  CeldasOcupadasCuadrado(15,6, 1) + CeldasOcupadasCuadrado(17,6, 1)+ CeldasOcupadasCuadrado(1,1, 8) +  CeldasOcupadasCuadrado(18,1, 8) + CeldasOcupadasCuadrado(1,18, 8) + CeldasOcupadasCuadrado(17,17, 5) + CeldasOcupadasSolo(9,18, 8, False) + CeldasOcupadasSolo(18,10, 8, True))
print(TotalCeldasOcupadas)
print("PRIMER PRINT")
wb = Workbook()
ws = wb.active
ColorFondo = "FFFFFF"
ColorQR = "000000"
StringTest = "01110111"
Green = "00FF00"
Blue = "0000FF"
ws.column_dimensions["B"].width = 3
PintarBase(ColorFondo,28)
PintarBaseQR(ColorQR)
AlternarPintar(ColorQR,"FF0000", 9, "Fila", 9, "A", 7,2)
AlternarPintar(ColorQR,"FF0000", 9, "Columna", 10, "G", 7,2)
Cadena = input("Ingrese una cadena de texto: ")
StringBinaria, Longitud = CadenaBinario(Cadena)

print(f"Sin funcion{StringBinaria}")
print(type(StringBinaria))
print(f"Funcion limpiar string {LimpiarString((StringBinaria))}")
NuevaString = LimpiarString((StringBinaria))
print(NuevaString)
print(Longitud)
LongitudBinaria = NumBinario(Longitud)
print(type(LongitudBinaria))
LongitudBinaria = LimpiarString(LongitudBinaria)
sin_espacios = str(LongitudBinaria).replace(" ", "")
print(sin_espacios)
PintarDiagonal(12,25, Green, "FF0000",NuevaString[0:8] )
PintarDiagonal(16,25, "0F0F0F", "00FF00",NuevaString[9:17] )
PintarDiagonal(20,25, Green, Blue,StringTest )
PintarDiagonal(24,25, Green, "0000FF",sin_espacios )
CeldasZigZag = []
for Fila in reversed(range(1, 26)):
    for Columna in reversed(range(1, 26)):
        CeldasZigZag.append(abecedario[Columna] + str(Fila + 1))


wb.save("Test.xlsx")