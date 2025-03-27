from openpyxl import Workbook
from openpyxl.styles import PatternFill
import re

def BinarioFunc(Char):
    ListaBinario = list("00000000")
    Base = 2
    if (Char % 2) == 0:
        ListaBinario[7] = "0"
    else:
        ListaBinario[7] = "1"
        Char += -1
    for i in range(0, 7):
        if int((Char / (Base ** (7 - i)))) == 1:
            ListaBinario[i] = "1"
            Char = (Char - (Base ** (7 - i)))
    return ListaBinario


ListaCompleta = str(input("Ingresa una palabra para transformarla a binario: "))
print(ListaCompleta)
LongitudInput = int(len(ListaCompleta))
print(f"La longitud es: {LongitudInput}")
MatrizBinaria = []
for i in range(0,LongitudInput):
    Aux = ord(ListaCompleta[i])
    print(str(BinarioFunc(Aux)))
    print(BinarioFunc(Aux))
    MatrizBinaria.append(str(BinarioFunc(Aux)))


print(MatrizBinaria)
ListaBinaria = list(MatrizBinaria)
print(ListaBinaria)
MatrizLimpia =[re.sub(r"[\[\], ]", "", elem) for elem in MatrizBinaria]
print(str(MatrizLimpia))
String = str(MatrizLimpia)
String = "".join(filter(lambda letra: letra != "[", String))
String = "".join(filter(lambda letra: letra != "]", String))
String = "".join(filter(lambda letra: letra != '"', String))
String = "".join(filter(lambda letra: letra != "'", String))
String = "".join(filter(lambda letra: letra != ',', String))
print(String)

# Crear un nuevo libro de Excel y seleccionar la hoja activa
wb = Workbook()
ws = wb.active

# Aplicar color a la celda B2

abecedario = ["A","B","C","D","E","F","G","H","I","J","K","L","M","N","O","P","Q","R","S","T","U","V","W","X","Y","Z","AA","AB","AC","AD","AE","AF","AG","AH","AI","AJ","AK","AL", "AM","AN" , "AO", "AP", "AQ", "AR", "AS", "AT", "AU","AV", "AW", "AX", "AY", "AZ"]
columna = "B"
fila = 1
color = "000000"
ws.column_dimensions["B"].width = 3
aux2= 0

for i in range(0, len(String)):
    print(f"Estoy coloreando esta celda {abecedario[aux2] + str(fila)} con este número: {String[i]}")
    if (String[i] == " "):
        aux2 += 1
        ws.column_dimensions[abecedario[aux2]].width = 3
        fila = fila - 9

    else:
        aux = int(String[i])
        if  aux == 0:
            color = "FFFFFF"
        elif aux == 1: color = "000000"
    ws.column_dimensions[abecedario[aux2]].width = 3
    if not (String[i] == " "):
        fill = PatternFill(start_color=color, end_color=color, fill_type="solid")
        ws[abecedario[aux2] + str(fila)].fill = fill
    fila = fila + 1




# Guardar el archivo
wb.save("QRGenerado.xlsx")
