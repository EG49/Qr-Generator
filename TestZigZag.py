import openpyxl
from openpyxl.styles import PatternFill


fila = range(1,27)
abecedario = ["A", "B", "C", "D", "E", "F", "G", "H", "I", "J", "K", "L", "M", "N", "O", "P", "Q", "R", "S", "T",
                  "U", "V", "W", "X", "Y", "Z", "AA"]
def PintarBase(Color, Rango):
    fill = PatternFill(start_color=Color, end_color=Color, fill_type="solid")
    for i in range(1, Rango):
        for i2 in range(1, Rango):
            celda = abecedario[i - 1]  + str(i2)
            ws[celda].fill = fill
            ws.column_dimensions[abecedario[i - 1]].width = 3

def PintarRango(Color, Rango, Linea, PuntoInicio,Columna, Fila ):
    fill = PatternFill(start_color=Color, end_color=Color, fill_type="solid")
    if Linea == "Fila":
        for i in range(1, Rango):
            celda = abecedario[i + PuntoInicio - 1] + str(Fila)
            ws[celda].fill = fill
            ws.column_dimensions[abecedario[i + PuntoInicio - 1]].width = 3
    else:
        for i in range(1, Rango):
            celda = Columna + str(i + PuntoInicio)
            ws[celda].fill = fill
            ws.column_dimensions[abecedario[i - 1]].width = 3

def PintarCelda(Celda, Color):
    fill = PatternFill(start_color=Color, end_color=Color, fill_type="solid")
    ws[Celda].fill = fill

def AlternarPintar(ColorQR, ColorFondo, Rango, Linea, PuntoInicio, Columna, Fila, NSaltos ):
    fill = PatternFill(start_color=ColorQR, end_color=ColorQR, fill_type="solid")
    if Linea == "Fila":
        for aux in range(0, Rango , NSaltos):
            try:
                PintarCelda(abecedario[aux + PuntoInicio] + str(Fila), ColorQR)

                aux = aux + PuntoInicio
                print("El contador i es " + str(aux))

            except: print("list index out of range")


    else:
        for aux in range(0, Rango, NSaltos):
            try:
                PintarCelda(Columna+ str(aux + PuntoInicio), ColorQR)

                aux = aux + PuntoInicio

            except:
                print("list index out of range")


def PintarBaseQR(ColorQR):
    PintarRango(ColorQR, 8, "Fila", 1, "A", 2)
    PintarRango(ColorQR, 8, "Columna", 1, "B", 2)
    PintarRango(ColorQR, 8, "Fila", 1, "A", 8)
    PintarRango(ColorQR, 8, "Columna", 19, "H", 20)
    PintarRango(ColorQR, 8, "Fila", 1, "A", 20)
    PintarRango(ColorQR, 8, "Columna", 19, "B", 26)
    PintarRango(ColorQR, 8, "Fila", 1, "A", 26)
    PintarRango(ColorQR, 8, "Columna", 1, "H", 2)
    PintarRango(ColorQR, 8, "Columna", 1, "T", 20)
    PintarRango(ColorQR, 8, "Fila", 19, "T", 2)
    PintarRango(ColorQR, 8, "Columna", 1, "Z", 26)
    PintarRango(ColorQR, 8, "Fila", 19, "Z", 8)
    PintarRango(ColorQR, 4, "Fila", 3, "A", 4)
    PintarRango(ColorQR, 4, "Fila", 3, "A", 5)
    PintarRango(ColorQR, 4, "Fila", 3, "A", 6)
    PintarRango(ColorQR, 4, "Fila", 3, "A", 22)
    PintarRango(ColorQR, 4, "Fila", 3, "A", 23)
    PintarRango(ColorQR, 4, "Fila", 3, "A", 24)
    PintarRango(ColorQR, 4, "Fila", 21, "A", 4)
    PintarRango(ColorQR, 4, "Fila", 21, "A", 5)
    PintarRango(ColorQR, 4, "Fila", 21, "A", 6)
    PintarRango(ColorQR, 6, "Fila", 17, "A", 18)
    PintarRango(ColorQR, 6, "Columna", 17, "R", 2)
    PintarRango(ColorQR, 6, "Fila", 17, "A", 22)
    PintarRango(ColorQR, 6, "Columna", 17, "V", 20)
    PintarCelda("T20", ColorQR)
    PintarCelda("J19", ColorQR)


def PintarDiagonal(Fila, Columna, ColorQR, ColorFondo, Cadena):
    print("Empiezo a pintar diagonal")
    print(f"Longitud de la cadena es: {len(Cadena)} ")
    for i in range (0,len(Cadena)):
        print(f"Cadena {Cadena} con índice {i} es {Cadena[i]} ")
        if (Cadena[i] == "0"):
            Diagonal(ColorFondo, i, Columna, Fila)
        else:
            Diagonal(ColorQR, i , Columna, Fila)


def Diagonal(Color, Contador, Columna, Fila):
    if Contador == 0:
        Celda = abecedario[Columna] + str(Fila)
        PintarCelda(Celda, Color)
    elif Contador == 1:
        Celda = abecedario[Columna - 1] + str(Fila)
        PintarCelda(Celda, Color)
    elif Contador == 2:
        Celda = abecedario[Columna] + str(Fila - 1)
        PintarCelda(Celda, Color)
    elif Contador == 3:
        Celda = abecedario[Columna - 1] + str(Fila - 1)
        PintarCelda(Celda, Color)
    elif Contador == 4:
        Celda = abecedario[Columna] + str(Fila - 2)
        PintarCelda(Celda, Color)
    elif Contador == 5:
        Celda = abecedario[Columna - 1] + str(Fila - 2)
        PintarCelda(Celda, Color)
    elif Contador == 6:
        Celda = abecedario[Columna] + str(Fila - 3)
        PintarCelda(Celda, Color)
    elif Contador == 7:
        Celda = abecedario[Columna - 1] + str(Fila - 3)
        PintarCelda(Celda, Color)

# Crear libro
wb = openpyxl.Workbook()
ws = wb.active

# Colores
blanco = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
negro = PatternFill(start_color="000000", end_color="000000", fill_type="solid")



ColorFondo = "FFFFFF"
ColorQR = "000000"
StringTest = "01110111"
Green = "00FF00"
Blue = "0000FF"
ws.column_dimensions["B"].width = 3
PintarBase(ColorFondo,28)
PintarBaseQR(ColorQR)
AlternarPintar(ColorQR,"FF0000", 9, "Fila", 9, "A", 7,2)
AlternarPintar(ColorQR,"FF0000", 9, "Columna", 10, "G", 7,2)
# Cadena de bits binarios
# Todos los datos juntos: datos originales + códigos de corrección Reed-Solomon
data_total = [
    15, 104, 116, 116, 112, 115, 58, 47, 47, 116, 105, 110, 121, 117, 114, 108,
    46, 99, 111, 109, 47, 52, 104, 52, 104, 97, 99, 51, 114,
    28, 9, 195, 2, 10, 232, 246, 198, 72, 82
]

# Convertir cada byte a binario de 8 bits
binary_list = [format(byte, '08b') for byte in data_total]

# Concatenar todos los bits en una sola cadena
binary_string = ''.join(binary_list)


binary_data = (
    "00001111" "01101000" "01110100" "01110100" "01110000" "01110011" "00111010" "00101111" "00101111" "01110100"
    "01101001" "01101110" "01111001" "01110101" "01110010" "01101100" "00101110" "01100011" "01101111" "01101101"
    "00101111" "00110100" "01101000" "00110100" "01101000" "01100001" "01100011" "00110011" "01110010"
    "00011100" "00001001" "11000011" "00000010" "00001010" "11101000" "11110110" "11000110" "01001000" "01010010"
)

# Actualmente tenemos 312 bits (39 bytes). Para un QR versión 2 nivel Q se usan 44 bytes = 352 bits
# Debemos añadir 352 - 312 = 40 bits de terminación y relleno

# Paso 1: Añadir hasta 4 bits de terminación (relleno con ceros si no hay espacio para más)
terminator = "0000"
remaining = 352 - (len(binary_data) + len(terminator))

# Paso 2: Relleno hasta completar múltiplos de 8 bits (bytes)
padding_bits = ""
while (len(binary_data) + len(terminator) + len(padding_bits)) < 352:
    padding_bits += "1110110000010001"  # Se alternan estos dos bytes según el estándar

# Limitar a los bits necesarios exactos
total_bits = binary_data + terminator + padding_bits
total_bits = total_bits[:352]

# Verificar cantidad final
len(total_bits), total_bits

binarios = total_bits
indice_binario = 0

# Lista de celdas prohibidas (omitida aquí por brevedad, asume que sigue igual)
celdas_prohibidas = set([
    'L7', 'N7', 'P7', 'R7', 'G12', 'G14', 'G16', 'G18', 'B2', 'C2', 'D2', 'E2', 'F2', 'G2', 'H2', 'I2',
    'B3', 'C3', 'D3', 'E3', 'F3', 'G3', 'H3', 'I3', 'B4', 'C4', 'D4', 'E4', 'F4', 'G4', 'H4', 'I4',
    'B5', 'C5', 'D5', 'E5', 'F5', 'G5', 'H5', 'I5', 'B6', 'C6', 'D6', 'E6', 'F6', 'G6', 'H6', 'I6',
    'B7', 'C7', 'D7', 'E7', 'F7', 'G7', 'H7', 'I7', 'B8', 'C8', 'D8', 'E8', 'F8', 'G8', 'H8', 'I8',
    'B9', 'C9', 'D9', 'E9', 'F9', 'G9', 'H9', 'I9', 'B19', 'C19', 'D19', 'E19', 'F19', 'G19', 'H19', 'I19',
    'B20', 'C20', 'D20', 'E20', 'F20', 'G20', 'H20', 'I20', 'B21', 'C21', 'D21', 'E21', 'F21', 'G21', 'H21', 'I21',
    'B22', 'C22', 'D22', 'E22', 'F22', 'G22', 'H22', 'I22', 'B23', 'C23', 'D23', 'E23', 'F23', 'G23', 'H23', 'I23',
    'B24', 'C24', 'D24', 'E24', 'F24', 'G24', 'H24', 'I24', 'B25', 'C25', 'D25', 'E25', 'F25', 'G25', 'H25', 'I25',
    'B26', 'C26', 'D26', 'E26', 'F26', 'G26', 'H26', 'I26', 'S2', 'T2', 'U2', 'V2', 'W2', 'X2', 'Y2', 'Z2',
    'S3', 'T3', 'U3', 'V3', 'W3', 'X3', 'Y3', 'Z3', 'S4', 'T4', 'U4', 'V4', 'W4', 'X4', 'Y4', 'Z4',
    'S5', 'T5', 'U5', 'V5', 'W5', 'X5', 'Y5', 'Z5', 'S6', 'T6', 'U6', 'V6', 'W6', 'X6', 'Y6', 'Z6',
    'S7', 'T7', 'U7', 'V7', 'W7', 'X7', 'Y7', 'Z7', 'S8', 'T8', 'U8', 'V8', 'W8', 'X8', 'Y8', 'Z8',
    'S9', 'T9', 'U9', 'V9', 'W9', 'X9', 'Y9', 'Z9', 'R18', 'S18', 'T18', 'U18', 'V18', 'R19', 'S19', 'T19',
    'U19', 'V19', 'R20', 'S20', 'T20', 'U20', 'V20', 'R21', 'S21', 'T21', 'U21', 'V21', 'R22', 'S22', 'T22',
    'U22', 'V22', 'S10', 'T10', 'U10', 'V10', 'W10', 'X10', 'Y10', 'Z10', 'J19', 'J20', 'J21', 'J22',
    'J23', 'J24', 'J25', 'J26','B10', 'C10', 'D10', 'E10', 'F10', 'G10', 'H10', 'I10', 'J10', 'J2', 'J3', 'J4',
    'J5', 'J6', 'J7', 'J8', 'J9', 'J10',
]) # Usa la lista completa que diste antes

# Columnas desde Z hasta B
columnas = [chr(c) for c in range(ord('Z'), ord('B') - 1, -1)]
columnas_pares = [columnas[i:i+2] for i in range(0, len(columnas), 2)]

# Recorrido en zigzag con límite en B y fila 2
for par in columnas_pares:
    col1 = par[0]
    col2 = par[1] if len(par) == 2 else None

    if columnas.index(col1) % 2 == 0:
        filas = range(26, 1, -1)  # Desde 26 hasta 2
    else:
        filas = range(2, 27)      # Desde 2 hasta 26

    for fila in filas:
        for col in [col1, col2]:
            if col and indice_binario < len(binarios):
                celda = f"{col}{fila}"
                if celda in celdas_prohibidas:
                    continue  # Saltar esta celda

                color = negro if binarios[indice_binario] == '1' else blanco
                ws[celda].fill = color
                indice_binario += 1

# Guardar el archivo
wb.save("zigzag_b_hasta_fila2.xlsx")
